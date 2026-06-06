#!/usr/bin/env python3
"""enigma_export.py — turns computed quote data into Enigma-branded outputs:
   - internal styled .xlsx (full BUY/sell/profit columns, green section bars, live formulas)
   - client-facing .pdf  (4-col, ENIGMA wordmark, centered title block, footer, T&C page)

Designed to receive a payload the JS engine already produces:
  payload = {
    "header": {"client","event","job","version","date"},
    "sections": [ {"title","excl", "lines":[ {"item","qty","unitBuy","supplier","contact","note","confidence"} ]} ]
  }
This file also includes a tiny engine + sample data so it can be run standalone for a demo.
"""
import json, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

GREEN = "D7E4BD"   # section bars  (accent3 @ tint .6)
BLUE  = "B9CDE5"   # BUY / cost columns (accent1 @ tint .6)
FONT  = "Proxima Nova Regular"
MONEY = '#,##0.00'

# ---------- tiny engine (mirror of engine.js) so the demo reflects real numbers ----------
RATES = [
 ("fridge magnet","giveaway","piece",16,"tahadogifts","","high"),
 ("tote|bag","giveaway","piece",25,"tahadogifts","","low"),
 ("cutout|logo|letter","scenic","m2",370,"Bolt","","med"),
 ("step & repeat|media wall|wall","scenic","m2",370,"Bolt","","high"),
 ("photo booth|photobooth|360","activation","day",1500,"Diamond","","low"),
 ("host|hostess|promoter","entertainment","show",770,"JAM","James Mistry","high"),
 ("project manager","staffing","day",2500,"Enigma","","high"),
]
SECTION_TITLES = {"av":"PRODUCTION - AV & BRANDING","scenic":"SCENIC & FABRICATION",
 "activation":"ACTIVATIONS","giveaway":"GIVEAWAYS","entertainment":"ENTERTAINMENT & TALENT",
 "staffing":"STAFFING"}
SECTION_ORDER = ["av","scenic","activation","giveaway","entertainment","staffing"]
UNIT = {"mm":0.001,"cm":0.01,"m":1}

def classify(name):
    n=(name or "").lower()
    for kws,cat,drv,cost,sup,con,conf in RATES:
        for k in kws.split("|"):
            if k.strip() and k.strip() in n: return cat,drv,cost,sup,con,conf
    return "giveaway","piece",0,"","","low"

def build_payload(header, items):
    secs={}
    for it in items:
        cat,drv,cost,sup,con,conf=classify(it["item"])
        qty=float(it.get("qty") or 1)
        if drv=="m2":
            w=float(it.get("w") or 0); h=float(it.get("h") or 0); f=UNIT.get(it.get("unit","mm"),0.001)
            area=round(w*f*h*f,4) or 1; unit_buy=round(cost*area,2); qcol=qty
        elif drv=="day":
            unit_buy=cost; qcol=float(it.get("days") or qty)
        else:
            unit_buy=cost; qcol=qty
        secs.setdefault(cat,[]).append(dict(item=it["item"],qty=qcol,unitBuy=unit_buy,
            supplier=sup,contact=con,note="",confidence=conf))
    sections=[]
    for cat in SECTION_ORDER:
        if cat in secs:
            sections.append(dict(title=SECTION_TITLES[cat], excl=(cat=="staffing"), lines=secs[cat]))
    return dict(header=header, sections=sections)

# ---------- Enigma standard terms (their own template boilerplate) ----------
TC = {
 "TERMS & CONDITIONS":[
  "All items are on a rental basis and for a one time installation / dismantle, unless mentioned specifically. Elements and items proposed are subject to availability upon confirmation",
  "Final Costs will be based on final review and confirmation of all the requirements, detailed technical drawings, site visits and time available when the project is signed off by both parties",
  "The cost of the project will be incurred through the approved Purchase Order (PO). Any alterations to the requirements or scope will impact the initial costing and result in variations subject to mutual agreement",
  "A detailed project plan will be presented upon being awarded and will be updated when relevant, and in the circumstances of certain elements being not available - suitable replacement options will be suggested where possible based upon client approval",
  "Subject to venue h&s approvals and government jurisdictions, certain elements may need to be changed and subsequently modified in budget upon confirmation and approval",
  "Delays in approvals may affect the quality of the product and delivery time and the budget allocated for that line item, as such a deadline request may be placed in order to negate these effects",
  "Designs and Renders will be provided and up to 2 rounds of revisions are complimentary. Additional or new scope of work will be charged extra and based on agreed rates between the two parties",
 ],
 "BILLING AND PAYMENT":[
  "All costs based on final confirmation, detailed technical drawings and site visits. Terms of Payment will be as follows unless otherwise pre-agreed by the agency and client:",
  "50% - Advance Payment","50% - 30 Days Post Event",
 ],
 "CANCELLATION POLICY":[
  "If the client cancels or postpones the project after acceptance, then besides the below terms the agency reserves the right to charge a higher cancellation fee in the event the agency has paid 3rd party costs, manpower time and internal agency time which exceeds the % mentioned above and will be based on fair value and due diligence. The agency will also endeavour to reduce the charges of cancellation fee where possible for the client.",
  "A cancellation fee of 25% of the total quote is payable if the project is cancelled 1 month prior to the event",
  "A cancellation fee of 50% of the total quote is payable if the project is cancelled 2 weeks prior to the event",
  "A cancellation fee equalling to 100% of the total quote is payable if the project is cancelled 7 days before the event",
 ],
}
ADDRESS = "Unit 203, DSC Tower Dubai Studio City, Dubai, United Arab Emirates"

# =========================== INTERNAL XLSX ===========================
def build_xlsx(payload, path):
    h=payload["header"]; wb=Workbook(); ws=wb.active; ws.title="Estimated Budget"
    green=PatternFill("solid",fgColor=GREEN); blue=PatternFill("solid",fgColor=BLUE)
    bold=Font(name=FONT,size=12,bold=True); reg=Font(name=FONT,size=12)
    rt=Alignment(horizontal="right"); ctr=Alignment(horizontal="center")
    def C(r,c): return ws.cell(r,c)
    # header block
    C(6,1).value="ENIGMA"; C(7,1).value=h.get("event",""); 
    C(8,1).value=f"{h.get('version','Estimated Budget V1')} - {h.get('date','')}"
    C(9,1).value=f"Job Code - {h.get('job','')}"
    for r in range(6,10): C(r,1).font=bold
    # column headers row 13
    heads=["Item","Quantity","AED Unit BUY","AED Total BUY","AED Unit ","AED Total","Profit AED","Notes / Supplier"]
    for c,t in enumerate(heads,1):
        cell=C(13,c); cell.value=t; cell.font=bold
        if c in (3,4,7,8): cell.fill=blue
    r=15; sub_rows=[]; excl_rows=[]
    for sec in payload["sections"]:
        sr=r
        C(sr,1).value=sec["title"]; C(sr,2).value="SUB TOTAL"
        for c in range(1,9): C(sr,c).fill=green; C(sr,c).font=bold
        C(sr,2).alignment=ctr
        first=r+1
        for L in sec["lines"]:
            r+=1
            C(r,1).value=L["item"]; C(r,2).value=L["qty"]
            C(r,3).value=L["unitBuy"]; C(r,3).fill=blue; C(r,3).number_format=MONEY
            C(r,4).value=f"=C{r}*B{r}"; C(r,4).fill=blue; C(r,4).number_format=MONEY
            C(r,5).value=f"=(C{r}*1.25)"; C(r,5).number_format=MONEY
            C(r,6).value=f"=E{r}*B{r}"; C(r,6).number_format=MONEY
            C(r,7).value=f"=F{r}-D{r}"; C(r,7).number_format=MONEY
            C(r,8).value=(("["+L["confidence"].upper()+"] ") if L.get("confidence") else "")+(L.get("supplier") or "")+((" · "+L["contact"]) if L.get("contact") else "")
            for c in range(1,9): C(r,c).font=reg
        last=r
        C(sr,4).value=f"=SUM(D{first}:D{last})"; C(sr,4).number_format=MONEY
        C(sr,6).value=f"=SUM(F{first}:F{last})"; C(sr,6).number_format=MONEY
        C(sr,7).value=f"=SUM(G{first}:G{last})"; C(sr,7).number_format=MONEY
        C(sr,8).value=f"=IF(F{sr}=0,0,G{sr}/F{sr})"; C(sr,8).number_format='0%'
        sub_rows.append(sr); 
        if not sec["excl"]: excl_rows.append(sr)
        r+=2
    # management fee
    mr=r; C(mr,1).value="MANAGEMENT FEE @15% - EXCLUDING STAFFING"; C(mr,2).value=1
    for c in range(1,9): C(mr,c).fill=green; C(mr,c).font=bold
    C(mr,4).value=0; C(mr,4).number_format=MONEY
    C(mr,5).value="="+("+".join(f"F{s}" for s in excl_rows) or "0")+"*0.15"; C(mr,5).number_format=MONEY
    C(mr,6).value=f"=E{mr}*B{mr}"; C(mr,6).number_format=MONEY
    C(mr,7).value=f"=F{mr}-D{mr}"; C(mr,7).number_format=MONEY
    r+=2
    allr=sub_rows+[mr]; tr=r
    C(tr,1).value="Total"; C(tr,1).font=bold
    C(tr,4).value="="+"+".join(f"D{s}" for s in allr); C(tr,4).number_format=MONEY; C(tr,4).font=bold
    C(tr,6).value="="+"+".join(f"F{s}" for s in allr); C(tr,6).number_format=MONEY; C(tr,6).font=bold
    C(tr,7).value="="+"+".join(f"G{s}" for s in allr); C(tr,7).number_format=MONEY; C(tr,7).font=bold
    r+=1; vr=r; C(vr,1).value="VAT TOTAL at 5%"; C(vr,1).font=bold
    C(vr,6).value=f"=F{tr}*0.05"; C(vr,6).number_format=MONEY; C(vr,6).font=bold
    r+=1; C(r,1).value="Total inclusive of VAT"; C(r,1).font=bold
    C(r,6).value=f"=F{tr}+F{vr}"; C(r,6).number_format=MONEY; C(r,6).font=bold
    # T&C
    r+=3
    for hd,items in TC.items():
        C(r,1).value=hd; C(r,1).font=bold; r+=1
        for cl in items: C(r,1).value=cl; C(r,1).font=reg; r+=1
        r+=1
    # widths
    widths={1:81.2,2:14.5,3:19.7,4:16,5:17.3,6:15.7,7:18.7,8:57.3}
    for c,w in widths.items(): ws.column_dimensions[get_column_letter(c)].width=w
    ws.sheet_view.showGridLines=False
    wb.save(path)

# =========================== CLIENT PDF ===========================
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.platypus import (BaseDocTemplate, PageTemplate, Frame, Table, TableStyle,
    Paragraph, Spacer, PageBreak, Flowable)
from reportlab.lib.styles import ParagraphStyle

GREEN_C=colors.HexColor("#D7E4BD"); GREY=colors.HexColor("#8a8a8a")
LINE=colors.HexColor("#cfd4da"); DARK=colors.HexColor("#1c2430")

def aed(n): return f"{n:,.2f}"

def compute(payload):
    secs=[]; mgmt_base=0; total_buy=0; total_sell=0
    for sec in payload["sections"]:
        lines=[]; sub=0
        for L in sec["lines"]:
            us=round(L["unitBuy"]*1.25,2); ts=round(us*L["qty"],2)
            sub+=ts; total_sell+=ts; total_buy+=round(L["unitBuy"]*L["qty"],2)
            lines.append((L["item"],L["qty"],us,ts))
        secs.append((sec["title"],sub,lines)); 
        if not sec["excl"]: mgmt_base+=sub
    mgmt=round(mgmt_base*0.15,2); total=round(total_sell+mgmt,2)
    vat=round(total*0.05,2)
    return secs,mgmt,total,vat,round(total+vat,2)

class Wordmark(Flowable):
    def __init__(self,size=26): self.size=size; self.width=0; self.height=size+6
    def draw(self):
        c=self.canv; to=c.beginText(0,0); to.setFont("Helvetica-Bold",self.size)
        to.setFillColor(DARK); to.setCharSpace(self.size*0.28); to.textOut("ENIGMA"); c.drawText(to)

def build_pdf(payload, path):
    h=payload["header"]; secs,mgmt,total,vat,grand=compute(payload)
    doc=BaseDocTemplate(path,pagesize=A4,leftMargin=18*mm,rightMargin=18*mm,
                        topMargin=16*mm,bottomMargin=18*mm)
    frame=Frame(doc.leftMargin,doc.bottomMargin,doc.width,doc.height,id="f")
    def furniture(c,d):
        c.saveState(); c.setFont("Helvetica",8); c.setFillColor(GREY)
        c.drawCentredString(A4[0]/2,12*mm,ADDRESS); c.restoreState()
    doc.addPageTemplates([PageTemplate(id="all",frames=[frame],onPage=furniture)])

    PS=ParagraphStyle; body=PS("b",fontName="Helvetica",fontSize=8.2,leading=11,textColor=DARK)
    ctr=PS("c",parent=body,alignment=1,fontName="Helvetica-Bold",fontSize=9)
    note=PS("n",parent=body,alignment=1,textColor=GREY,fontSize=7.5)
    bar=PS("bar",fontName="Helvetica-Bold",fontSize=8.5,textColor=colors.white,leading=12)
    tch=PS("tch",fontName="Helvetica-Bold",fontSize=9,textColor=colors.white)
    story=[]
    story+=[Wordmark(26), Spacer(1,18*mm)]
    for t in ["ENIGMA",h.get("event",""),f"{h.get('version','Estimated Budget V1')} - {h.get('date','')}",f"Job Code - {h.get('job','')}"]:
        story.append(Paragraph(t,ctr))
    story.append(Spacer(1,3))
    story.append(Paragraph("Estimated quote — auto-derived from historical supplier rates. Subject to confirmation &amp; final scope.",note))
    story.append(Spacer(1,10))

    # table
    data=[["Item","Quantity","AED Unit","AED Total"]]; styles=[]
    ts=[("FONTNAME",(0,0),(-1,-1),"Helvetica"),("FONTSIZE",(0,0),(-1,-1),8.2),
        ("ALIGN",(1,0),(-1,-1),"CENTER"),("ALIGN",(3,0),(3,-1),"RIGHT"),("ALIGN",(2,0),(2,-1),"RIGHT"),
        ("ALIGN",(0,0),(0,-1),"LEFT"),("TEXTCOLOR",(0,0),(-1,-1),DARK),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("LINEBELOW",(0,0),(-1,0),0.8,DARK),
        ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
        ("LEFTPADDING",(0,0),(-1,-1),5),("RIGHTPADDING",(0,0),(-1,-1),5)]
    ri=1
    for title,sub,lines in secs:
        data.append([title,"SUB TOTAL","",aed(sub)])
        ts+=[("BACKGROUND",(0,ri),(-1,ri),GREEN_C),("FONTNAME",(0,ri),(-1,ri),"Helvetica-Bold"),
             ("ALIGN",(1,ri),(1,ri),"CENTER")]
        ri+=1
        for item,qty,us,tsell in lines:
            q=f"{qty:g}"
            data.append([item,q,aed(us),aed(tsell)])
            ts+=[("LINEBELOW",(0,ri),(-1,ri),0.4,LINE)]; ri+=1
        data.append(["","","",""]); ts+=[("TOPPADDING",(0,ri),(-1,ri),2),("BOTTOMPADDING",(0,ri),(-1,ri),2)]; ri+=1
    # mgmt fee
    data.append(["MANAGEMENT FEE @15% - EXCLUDING STAFFING","1",aed(mgmt),aed(mgmt)])
    ts+=[("BACKGROUND",(0,ri),(-1,ri),GREEN_C),("FONTNAME",(0,ri),(-1,ri),"Helvetica-Bold")]; ri+=1
    # totals
    for lab,val in [("Total",total),("VAT TOTAL at 5%",vat),("Total inclusive of VAT",grand)]:
        data.append([lab,"","",aed(val)])
        ts+=[("FONTNAME",(0,ri),(-1,ri),"Helvetica-Bold"),("LINEABOVE",(0,ri),(-1,ri),0.6,DARK)]; ri+=1
    tbl=Table(data,colWidths=[doc.width-3*30*mm,30*mm,30*mm,30*mm],repeatRows=1)
    tbl.setStyle(TableStyle(ts)); story.append(tbl)

    # ---- T&C page ----
    story.append(PageBreak())
    def black_bar(text):
        t=Table([[Paragraph(text,tch)]],colWidths=[doc.width])
        t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),DARK),("LEFTPADDING",(0,0),(-1,-1),6),
            ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4)])); return t
    for hd,items in TC.items():
        story.append(black_bar(hd)); story.append(Spacer(1,4))
        for cl in items:
            story.append(Paragraph(cl,body))
            story.append(Spacer(1,2))
            story.append(Table([[""]],colWidths=[doc.width],style=[("LINEBELOW",(0,0),(-1,-1),0.4,LINE)]))
            story.append(Spacer(1,4))
        story.append(Spacer(1,6))
    doc.build(story)

# =========================== demo ===========================
if __name__=="__main__":
    items=[
     {"item":"Fridge Magnets - 7cm","qty":"250","unit":"mm","w":"70","h":"70"},
     {"item":"Cutout Logo (freestanding)","qty":"9","w":"2000","h":"1500","unit":"mm"},
     {"item":"Step & Repeat Media Wall","qty":"1","w":"3","h":"2.4","unit":"m"},
     {"item":"Photo Booth Activation","days":"3"},
     {"item":"Event Hostess","qty":"4","duration":"7"},
     {"item":"Branded Tote Bags","qty":"200"},
     {"item":"Project Manager","days":"2"},
    ]
    header={"client":"Property Finder","event":"Property Finder — Brand Activation Day",
            "job":"PF60500","version":"Estimated Budget V1",
            "date":datetime.date.today().strftime("%d/%m/%Y")}
    payload=build_payload(header,items)
    build_xlsx(payload,"/mnt/user-data/outputs/Sample_Estimated_Budget_INTERNAL.xlsx")
    build_pdf(payload,"/mnt/user-data/outputs/Sample_Estimated_Quote_CLIENT.pdf")
    print("done")
